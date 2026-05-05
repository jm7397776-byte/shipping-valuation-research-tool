#!/usr/bin/env python3
"""Audit and summarize the Red Sea shipping-shock DiD workbook.

The workbook is treated as the user-supplied vendor extract. This script does
not overwrite the raw sheets. It produces a compact JSON payload for the web app
and an enhanced audit workbook with regression result sheets.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path.home() / "Desktop" / "Red_Sea_DiD_Model_drafting.xlsx"
DEFAULT_JSON = Path("data/red_sea_shock_analysis.json")
DEFAULT_WORKBOOK = Path.home() / "Desktop" / "Red_Sea_DiD_Model_validated.xlsx"
DEFAULT_COMPLETED_WORKBOOK = Path.home() / "Desktop" / "Red_Sea_DiD_Model_completed.xlsx"
DEFAULT_STATA_DIR = Path.home() / "Desktop" / "red_sea_stata_package"


SOURCE_MAP = [
    {
        "item": "Company stock returns, close prices, market cap",
        "workbook_sheet": "Price_Daily_Raw",
        "workbook_source": "Refinitiv/LSEG Workspace in workbook Source_Note",
        "validation_source": "Cross-check with Bloomberg BDP/BDH, LSEG Workspace, or exchange close data",
        "status": "vendor",
        "note": "Workbook labels this source as Refinitiv/LSEG, not Bloomberg. If original pull was Bloomberg, keep the Bloomberg field/ticker log as an additional evidence sheet.",
    },
    {
        "item": "Market index return",
        "workbook_sheet": "Market_Index_Raw",
        "workbook_source": "Refinitiv/LSEG Workspace; S&P 500 in shown extract",
        "validation_source": "MSCI World/S&P index vendor export. Use one market benchmark consistently across the market model.",
        "status": "vendor",
        "note": "Return_Calc_Template states MSCI World, while Market_Index_Raw sample header shows S&P 500. This benchmark naming should be reconciled before final thesis use.",
    },
    {
        "item": "VIX",
        "workbook_sheet": "VIX_Raw",
        "workbook_source": "Refinitiv/LSEG Workspace",
        "validation_source": "Cboe VIX or FRED VIXCLS series",
        "status": "public cross-check available",
        "note": "FRED can be used for a reproducible VIX cross-check; daily timing and holidays still need alignment.",
    },
    {
        "item": "Brent crude",
        "workbook_sheet": "Freight_Context_Raw",
        "workbook_source": "Refinitiv LCOc1",
        "validation_source": "ICE Brent continuous futures via licensed vendor; FRED DCOILBRENTEU for spot cross-check",
        "status": "public cross-check available",
        "note": "Use futures or spot consistently; do not mix if interpreting as tradable market reaction.",
    },
    {
        "item": "BDI / BDTI / BCTI freight indices",
        "workbook_sheet": "Freight_Context_Raw",
        "workbook_source": "Clarksons SIN / Baltic Exchange index series in workbook note",
        "validation_source": "Baltic Exchange licensed data, Clarksons SIN, Bloomberg, LSEG",
        "status": "licensed preferred",
        "note": "Official Baltic time series are licensed benchmark data. Public webpages can validate definitions, not the full historical daily panel.",
    },
    {
        "item": "Company financials",
        "workbook_sheet": "Financials_Raw",
        "workbook_source": "Refinitiv/LSEG Workspace in workbook Source_Note/title",
        "validation_source": "Annual report, 20-F/10-K, exchange filing, SEC EDGAR for SEC filers",
        "status": "filing final",
        "note": "For thesis-final valuation multiples, replace vendor snapshot values with filing/audited values or document the vendor snapshot date.",
    },
]


CHECK_KO = {
    "Control group formula row reference": {
        "check_ko": "통제군 공식 행 참조 오류",
        "detail_ko": "Control 공식이 현재 행이 아니라 바로 위 행의 탱커/벌커 비중을 참조합니다. 엑셀을 다시 계산하면 통제군 판정이 틀어질 수 있습니다.",
        "fix_ko": "각 행에서 DryBulk_%는 같은 행 G열, Tanker_%는 같은 행 F열을 보도록 수정했습니다.",
    },
    "CAR VLOOKUP index range": {
        "check_ko": "CAR 시트 조회 범위 오류",
        "detail_ko": "CAR_Calc_Template의 VLOOKUP이 선택한 범위보다 더 먼 열 번호를 요청해 IFERROR 때문에 빈칸이 됩니다.",
        "fix_ko": "RIC 기준 INDEX/MATCH 공식으로 Alpha, Beta, R2, N, Size, Leverage, Cash 값을 가져오게 수정했습니다.",
    },
    "Excel compatibility": {
        "check_ko": "엑셀 버전 호환성",
        "detail_ko": "일부 공식에 XLOOKUP 흔적이 있어 구버전 Excel에서는 #NAME? 오류가 날 수 있습니다.",
        "fix_ko": "완성본에서는 핵심 조회 공식을 INDEX/MATCH 중심으로 바꿨습니다.",
    },
    "Source label consistency": {
        "check_ko": "원자료 출처 표기 불일치",
        "detail_ko": "사용자는 Bloomberg 추출을 말했지만 엑셀 내부에는 Refinitiv/LSEG Workspace 표기가 있습니다.",
        "fix_ko": "최종 논문 전 Bloomberg 추출 로그를 붙이거나, Refinitiv/LSEG 출처로 통일해야 합니다.",
    },
    "Regression panel firm coverage": {
        "check_ko": "회귀 패널 회사 커버리지 차이",
        "detail_ko": "포함 회사 수보다 DiD 패널에 실제 들어간 회사 수가 적습니다. 가격 데이터 부족, 거래정지, 티커 불일치 가능성이 있습니다.",
        "fix_ko": "제외 회사와 제외 사유를 부록 표로 남기세요.",
    },
}


def add_korean_check_fields(issue: dict[str, Any]) -> dict[str, Any]:
    ko = CHECK_KO.get(issue.get("check"), {})
    issue["check_ko"] = ko.get("check_ko", issue.get("check", ""))
    issue["detail_ko"] = ko.get("detail_ko", issue.get("detail", ""))
    issue["fix_ko"] = ko.get("fix_ko", issue.get("fix", ""))
    return issue


def clean_col(value: Any) -> str:
    return str(value).replace("\n", " ").replace("×", "x").replace("−", "-").strip()


CANONICAL_COLUMNS = {
    "TreatxPost": "Treat_Post",
    "Size (ln MCap)": "Size_ln_MCap",
    "Size ln MCap": "Size_ln_MCap",
    "Cash/ Assets": "Cash_Assets",
    "Cash / Assets": "Cash_Assets",
    "Cash Assets": "Cash_Assets",
    "Mkt_Return (MSCI)": "Market_Return",
    "Oil_Return (Brent)": "Oil_Return",
    "Stock Return": "Return",
    "AR (Abnormal)": "AR",
    "Expected Return": "Expected_Return",
    "Est_Flag [-250,-30]": "Est_Flag",
    "CAR [-1,+5]": "CAR_5",
    "CAR [-1,+10]": "CAR_10",
    "CAR [-1,+20]": "CAR_20",
    "Pre-CAR [-20,-2]": "Pre_CAR",
    "AR_Treat (avg)": "AR_Treat",
    "AR_Ctrl (avg)": "AR_Ctrl",
    "AR_Diff (T-C)": "AR_Diff",
    "CAR_Treat (cumulative)": "CAR_Treat",
    "CAR_Ctrl (cumulative)": "CAR_Ctrl",
    "CAR_Diff (T-C)": "CAR_Diff",
    "R²": "R2",
    "Beta (Refinitiv)": "Beta_Refinitiv",
}


def read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, header=2)
    df.columns = [CANONICAL_COLUMNS.get(clean_col(col), clean_col(col)) for col in df.columns]
    return df


def to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def finite_rows(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in cols:
        out[col] = to_num(out[col])
    return out.dropna(subset=cols)


def normal_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def pvalue_from_t(t_stat: float) -> float | None:
    if not math.isfinite(t_stat):
        return None
    return 2.0 * (1.0 - normal_cdf(abs(t_stat)))


def ols_hc1(df: pd.DataFrame, y_col: str, x_cols: list[str], name: str) -> dict[str, Any]:
    cols = [y_col, *x_cols]
    clean = finite_rows(df, cols)
    if len(clean) <= len(x_cols) + 2:
        return {
            "name": name,
            "outcome": y_col,
            "n": int(len(clean)),
            "r2": None,
            "coefficients": [],
            "note": "Insufficient observations after dropping missing values.",
        }

    y = clean[y_col].to_numpy(dtype=float)
    x = clean[x_cols].to_numpy(dtype=float)
    x = np.column_stack([np.ones(len(x)), x])
    labels = ["Intercept", *x_cols]
    xtx_inv = np.linalg.pinv(x.T @ x)
    beta = xtx_inv @ x.T @ y
    fitted = x @ beta
    resid = y - fitted
    k = x.shape[1]
    n = x.shape[0]
    meat = x.T @ (resid[:, None] ** 2 * x)
    vcov = xtx_inv @ meat @ xtx_inv * (n / max(n - k, 1))
    se = np.sqrt(np.maximum(np.diag(vcov), 0))
    t_stats = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se > 0)
    ss_res = float(np.sum(resid**2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot else None

    return {
        "name": name,
        "outcome": y_col,
        "n": int(n),
        "r2": r2,
        "coefficients": [
            {
                "term": labels[i],
                "coef": float(beta[i]),
                "se_hc1": float(se[i]) if math.isfinite(se[i]) else None,
                "t": float(t_stats[i]) if math.isfinite(t_stats[i]) else None,
                "p_approx": pvalue_from_t(float(t_stats[i])) if math.isfinite(t_stats[i]) else None,
            }
            for i in range(len(labels))
        ],
        "note": "OLS with HC1 heteroskedasticity-robust standard errors; p-values use normal approximation.",
    }


def col_letter_to_num(col: str) -> int:
    value = 0
    for char in col:
        value = value * 26 + (ord(char.upper()) - 64)
    return value


def audit_formulas(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    issues: list[dict[str, Any]] = []

    ws = wb["Firm_Master"]
    bad_control = []
    for row in range(4, ws.max_row + 1):
        formula = ws.cell(row, 9).value
        if isinstance(formula, str) and formula.startswith("="):
            expected = f"G{row}"
            expected_f = f"F{row}"
            if expected not in formula or expected_f not in formula:
                bad_control.append({"cell": f"Firm_Master!I{row}", "formula": formula})
    if bad_control:
        issues.append(
            {
                "severity": "High",
                "check": "Control group formula row reference",
                "status": "Fail",
                "affected": len(bad_control),
                "detail": "Control formulas reference the previous row in many cells, so cached values may not recalculate correctly.",
                "examples": bad_control[:5],
                "fix": "Use =IF(AND(G{row}>=70,F{row}<=20),1,0) for each Firm_Master row.",
            }
        )

    ws = wb["CAR_Calc_Template"]
    vlookup_bad = []
    pattern = re.compile(r"VLOOKUP\([^,]+,\s*Firm_Master!\$([A-Z]+):\$([A-Z]+),\s*(\d+)", re.I)
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            formula = ws.cell(row, col).value
            if not (isinstance(formula, str) and "VLOOKUP" in formula.upper()):
                continue
            match = pattern.search(formula)
            if not match:
                continue
            start, end, idx = match.groups()
            width = col_letter_to_num(end) - col_letter_to_num(start) + 1
            if int(idx) > width:
                vlookup_bad.append(
                    {
                        "cell": f"CAR_Calc_Template!{get_column_letter(col)}{row}",
                        "formula": formula[:140],
                        "range_width": width,
                        "requested_index": int(idx),
                    }
                )
    if vlookup_bad:
        issues.append(
            {
                "severity": "High",
                "check": "CAR VLOOKUP index range",
                "status": "Fail",
                "affected": len(vlookup_bad),
                "detail": "CAR formulas request columns outside the selected Firm_Master range, which returns blanks under IFERROR.",
                "examples": vlookup_bad[:5],
                "fix": "Replace with XLOOKUP by RIC to the exact Alpha/Beta/R2/N_Est/Size/Leverage columns or widen the lookup range.",
            }
        )

    if any("_xlfn.XLOOKUP" in str(cell.value) for cell in wb["Firm_Master"].iter_rows(min_row=4, max_row=min(ws.max_row, 12), values_only=False) for cell in cell):
        issues.append(
            {
                "severity": "Medium",
                "check": "Excel compatibility",
                "status": "Review",
                "affected": "Firm_Master formulas",
                "detail": "_xlfn.XLOOKUP appears in formulas. Excel versions without XLOOKUP support may show #NAME? even when cached values display.",
                "examples": [],
                "fix": "Use INDEX/MATCH for maximum compatibility or require Excel 365.",
            }
        )

    return issues


def workbook_audit(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    xl = pd.ExcelFile(path)
    fm = read_sheet(path, "Firm_Master")
    price = read_sheet(path, "Price_Daily_Raw")
    fin = read_sheet(path, "Financials_Raw")
    market = read_sheet(path, "Market_Index_Raw")
    vix = read_sheet(path, "VIX_Raw")
    freight = read_sheet(path, "Freight_Context_Raw")
    did = read_sheet(path, "DiD_Panel_Template")
    car = read_sheet(path, "CAR_Calc_Template")

    checks: list[dict[str, Any]] = []
    checks.extend(audit_formulas(path))

    if "Source_Note" in price:
        price_sources = sorted(set(price["Source_Note"].dropna().astype(str)))
    else:
        price_sources = []
    source_mismatch = any("Refinitiv" in item or "LSEG" in item for item in price_sources)
    checks.append(
        {
            "severity": "Medium" if source_mismatch else "Review",
            "check": "Source label consistency",
            "status": "Review",
            "affected": "Workbook-level source notes",
            "detail": "The attached workbook labels price/financial/index sheets as Refinitiv/LSEG Workspace. The user mentioned Bloomberg, so the source trail should be reconciled.",
            "examples": price_sources[:5],
            "fix": "Add a Bloomberg extraction log sheet if Bloomberg was the real source, or retain Refinitiv/LSEG as the cited source.",
        }
    )

    price["Date"] = pd.to_datetime(price["Date"], errors="coerce")
    did["Date"] = pd.to_datetime(did["Date"], errors="coerce")
    duplicate_price_keys = int(price.duplicated(subset=["Date", "RIC"]).sum()) if {"Date", "RIC"} <= set(price.columns) else 0
    price_missing_return = int(to_num(price.get("Total_Return", pd.Series(dtype=float))).isna().sum())
    price_abs_median = float(to_num(price.get("Total_Return", pd.Series(dtype=float))).abs().median())
    return_scale_note = "Price_Daily_Raw Total_Return appears to be percent units; Return_Calc_Template uses decimals." if price_abs_median > 0.5 else "Return scale appears decimal."

    required_fin = ["Market_Cap", "Total_Assets", "Total_Debt", "Cash", "Book_Equity", "Revenue", "EBITDA"]
    missing_fin = {
        col: int(to_num(fin[col]).isna().sum())
        for col in required_fin
        if col in fin.columns
    }
    did_required = ["AR", "Treat", "Post", "Treat_Post", "Market_Return", "Oil_Return", "VIX_Close"]
    did_missing = {
        col: int(to_num(did[col]).isna().sum())
        for col in did_required
        if col in did.columns
    }

    did_firms = int(did["RIC"].nunique()) if "RIC" in did else 0
    included_firms = int(fm.loc[to_num(fm.get("Include", pd.Series(dtype=float))).eq(1), "RIC"].nunique()) if "Include" in fm else 0
    if did_firms < included_firms:
        checks.append(
            {
                "severity": "Medium",
                "check": "Regression panel firm coverage",
                "status": "Review",
                "affected": f"{did_firms} panel firms vs {included_firms} included firms",
                "detail": "Some included firms do not appear in the DiD panel, likely due to missing/illiquid return histories.",
                "examples": [],
                "fix": "Keep a firm exclusion log with reason: no price data, insufficient estimation window, suspended/illiquid, or ticker mismatch.",
            }
        )

    checks = [add_korean_check_fields(item) for item in checks]

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_workbook": path.name,
        "sheets": xl.sheet_names,
        "firm_count": int(len(fm.dropna(subset=["RIC"]))),
        "included_firms": included_firms,
        "tanker_firms": int(to_num(fm.get("Treatment", pd.Series(dtype=float))).eq(1).sum()),
        "control_firms": int(to_num(fm.get("Control", pd.Series(dtype=float))).eq(1).sum()),
        "price_observations": int(len(price.dropna(subset=["RIC"]))),
        "price_firms": int(price["RIC"].nunique()),
        "price_date_min": str(price["Date"].min().date()) if price["Date"].notna().any() else None,
        "price_date_max": str(price["Date"].max().date()) if price["Date"].notna().any() else None,
        "duplicate_price_keys": duplicate_price_keys,
        "price_missing_return": price_missing_return,
        "price_return_scale_note": return_scale_note,
        "financial_rows": int(len(fin.dropna(subset=["RIC"]))),
        "financial_missing_counts": missing_fin,
        "market_rows": int(len(market)),
        "vix_rows": int(len(vix)),
        "freight_rows": int(len(freight)),
        "did_rows": int(len(did.dropna(subset=["RIC"]))),
        "did_firms": did_firms,
        "did_event_min": int(to_num(did["EventDay"]).min()) if "EventDay" in did else None,
        "did_event_max": int(to_num(did["EventDay"]).max()) if "EventDay" in did else None,
        "did_missing_counts": did_missing,
        "car_rows": int(len(car.dropna(subset=["RIC"]))),
        "source_notes": {
            "price": price_sources,
            "financials": ["Refinitiv/LSEG Workspace per sheet title"],
            "freight": ["Refinitiv LCOc1 + Clarksons SIN per sheet note"],
        },
    }

    return summary, checks


def run_models(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    did = read_sheet(path, "DiD_Panel_Template")
    did["Treat_Post"] = to_num(did.get("Treat_Post", did.get("TreatxPost", pd.Series(dtype=float))))
    if "Include" in did.columns:
        did = did[to_num(did["Include"]).eq(1)]

    models = [
        ols_hc1(did, "AR", ["Treat", "Post", "Treat_Post"], "Base DiD: AR ~ Treat + Post + Treat×Post"),
        ols_hc1(
            did,
            "AR",
            ["Treat", "Post", "Treat_Post", "Size_ln_MCap", "Leverage", "Market_Return", "Oil_Return", "VIX_Close"],
            "Controlled DiD: + size/leverage/market/oil/VIX",
        ),
        ols_hc1(
            did,
            "AR",
            ["Treat", "Post", "Treat_Post", "BDTI_Return", "BCTI_Return", "BDI_Return"],
            "Freight sensitivity: + BDTI/BCTI/BDI returns",
        ),
        ols_hc1(
            did,
            "Return",
            ["Treat", "Post", "Treat_Post", "Market_Return", "Oil_Return", "VIX_Close"],
            "Raw return DiD: stock return outcome",
        ),
    ]

    event = read_sheet(path, "Event_Study_Path")
    event_rows = []
    for _, row in event.dropna(subset=["EventDay"]).iterrows():
        event_rows.append(
            {
                "event_day": int(row["EventDay"]),
                "ar_treat": none_or_float(row.get("AR_Treat")),
                "ar_control": none_or_float(row.get("AR_Ctrl")),
                "ar_diff": none_or_float(row.get("AR_Diff")),
                "car_treat": none_or_float(row.get("CAR_Treat")),
                "car_control": none_or_float(row.get("CAR_Ctrl")),
                "car_diff": none_or_float(row.get("CAR_Diff")),
                "n_treat": none_or_int(row.get("N_Treat")),
                "n_control": none_or_int(row.get("N_Ctrl")),
            }
        )

    car = read_sheet(path, "CAR_Calc_Template")
    if "Include" in car.columns:
        car = car[to_num(car["Include"]).eq(1)]
    car_summary = []
    for metric, label in [("CAR_5", "CAR [-1,+5]"), ("CAR_10", "CAR [-1,+10]"), ("CAR_20", "CAR [-1,+20]")]:
        clean = finite_rows(car, [metric, "Treat"])
        tanker = clean[to_num(clean["Treat"]).eq(1)][metric]
        control = clean[to_num(clean["Treat"]).eq(0)][metric]
        car_summary.append(
            {
                "metric": label,
                "tanker_avg": none_or_float(tanker.mean()),
                "control_avg": none_or_float(control.mean()),
                "difference": none_or_float(tanker.mean() - control.mean()),
                "n_tanker": int(tanker.count()),
                "n_control": int(control.count()),
            }
        )

    return models, event_rows, car_summary


def nearest_series_value(df: pd.DataFrame, target_date: pd.Timestamp, value_col: str) -> float | None:
    if target_date is pd.NaT or value_col not in df.columns:
        return None
    dated = df.dropna(subset=["Date"]).sort_values("Date")
    if dated.empty:
        return None
    before = dated[dated["Date"] <= target_date]
    if before.empty:
        before = dated
    return none_or_float(before.iloc[-1].get(value_col))


def build_valuation_event_panel(path: Path) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    """Reconstruct an event-window valuation panel from the workbook.

    This is not treated as a Bloomberg-grade final ledger. It is a transparent
    bridge: daily close prices from the workbook plus the workbook's
    report-date market cap, debt, cash, book equity, and EBITDA snapshot.
    """
    price = read_sheet(path, "Price_Daily_Raw")
    fin = read_sheet(path, "Financials_Raw")
    fm = read_sheet(path, "Firm_Master")
    did = read_sheet(path, "DiD_Panel_Template")

    price["Date"] = pd.to_datetime(price["Date"], errors="coerce")
    fin["Report_Date"] = pd.to_datetime(fin.get("Report_Date"), errors="coerce")
    did["Date"] = pd.to_datetime(did["Date"], errors="coerce")
    did["Treat_Post"] = to_num(did.get("Treat_Post", did.get("TreatxPost", pd.Series(dtype=float))))
    if "Include" in did.columns:
        did = did[to_num(did["Include"]).eq(1)].copy()
    if "Include" in fm.columns:
        fm = fm[to_num(fm["Include"]).eq(1)].copy()

    firm_meta = fm.set_index("RIC", drop=False)
    fin_meta = fin.set_index("RIC", drop=False)
    did_key = did[["Date", "EventDay", "RIC", "Treat", "Post", "Treat_Post", "Include"]].dropna(subset=["Date", "RIC"])
    merged = did_key.merge(
        price[["Date", "RIC", "Company_Name", "Close_Price", "Market_Cap", "Source_Note"]],
        on=["Date", "RIC"],
        how="left",
    )

    rows: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    for ric, firm_rows in merged.groupby("RIC", dropna=True):
        if ric not in fin_meta.index:
            missing.append({"RIC": ric, "reason": "Financials_Raw row missing"})
            continue
        fin_row = fin_meta.loc[ric]
        if isinstance(fin_row, pd.DataFrame):
            fin_row = fin_row.iloc[0]
        firm_row = firm_meta.loc[ric] if ric in firm_meta.index else None
        if isinstance(firm_row, pd.DataFrame):
            firm_row = firm_row.iloc[0]

        report_mcap = none_or_float(fin_row.get("Market_Cap"))
        report_date = pd.to_datetime(fin_row.get("Report_Date"), errors="coerce")
        price_rows = price[price["RIC"].eq(ric)].copy()
        base_close = nearest_series_value(price_rows, report_date, "Close_Price")
        implied_shares = report_mcap / base_close if report_mcap and base_close and base_close > 0 else None
        debt = none_or_float(fin_row.get("Total_Debt")) or 0.0
        cash = none_or_float(fin_row.get("Cash")) or 0.0
        ebitda = none_or_float(fin_row.get("EBITDA"))
        book_equity = none_or_float(fin_row.get("Book_Equity"))
        if implied_shares is None:
            missing.append({"RIC": ric, "reason": "Cannot infer shares from report-date market cap and close price"})

        for _, row in firm_rows.iterrows():
            close = none_or_float(row.get("Close_Price"))
            market_cap_reported = none_or_float(row.get("Market_Cap"))
            market_cap_reconstructed = market_cap_reported
            source_method = "reported daily market cap"
            if market_cap_reconstructed is None and close is not None and implied_shares is not None:
                market_cap_reconstructed = close * implied_shares
                source_method = "close price x implied shares from report-date market cap"
            enterprise_value = (
                market_cap_reconstructed + debt - cash
                if market_cap_reconstructed is not None
                else None
            )
            rows.append(
                {
                    "Date": row["Date"].date().isoformat() if pd.notna(row["Date"]) else None,
                    "EventDay": none_or_int(row.get("EventDay")),
                    "RIC": ric,
                    "Company_Name": row.get("Company_Name") or (firm_row.get("Company_Name") if firm_row is not None else ""),
                    "Group": "Tanker" if none_or_float(row.get("Treat")) == 1 else "Control",
                    "Treat": none_or_int(row.get("Treat")),
                    "Post": none_or_int(row.get("Post")),
                    "Close_Price": close,
                    "Market_Cap_Event": market_cap_reconstructed,
                    "Enterprise_Value_Event": enterprise_value,
                    "EV_EBITDA_Event": enterprise_value / ebitda if enterprise_value is not None and ebitda and ebitda > 0 else None,
                    "P_B_Event": market_cap_reconstructed / book_equity if market_cap_reconstructed is not None and book_equity and book_equity > 0 else None,
                    "Report_Date": report_date.date().isoformat() if pd.notna(report_date) else None,
                    "Report_Market_Cap": report_mcap,
                    "Total_Debt": debt,
                    "Cash": cash,
                    "EBITDA": ebitda,
                    "Book_Equity": book_equity,
                    "Implied_Shares": implied_shares,
                    "Source_Method": source_method,
                    "Source_Grade": "B-reconstructed vendor snapshot",
                    "Final_Use": "검증용 보조 패널. 논문 최종 밸류에이션 반응은 Bloomberg/LSEG/Refinitiv/공시 원장으로 교체.",
                }
            )

    panel = pd.DataFrame(rows)
    summary_rows: list[dict[str, Any]] = []
    if not panel.empty:
        metric_cols = ["Market_Cap_Event", "Enterprise_Value_Event", "EV_EBITDA_Event", "P_B_Event"]
        firm_changes = []
        for (ric, group), firm_panel in panel.groupby(["RIC", "Group"]):
            pre = firm_panel[(to_num(firm_panel["EventDay"]) >= -20) & (to_num(firm_panel["EventDay"]) <= -1)]
            post = firm_panel[(to_num(firm_panel["EventDay"]) >= 0) & (to_num(firm_panel["EventDay"]) <= 20)]
            if pre.empty or post.empty:
                continue
            item: dict[str, Any] = {"RIC": ric, "Group": group}
            for metric in metric_cols:
                pre_mean = none_or_float(to_num(pre[metric]).mean())
                post_mean = none_or_float(to_num(post[metric]).mean())
                item[f"{metric}_PreMean"] = pre_mean
                item[f"{metric}_PostMean"] = post_mean
                if pre_mean and post_mean and metric in {"Market_Cap_Event", "Enterprise_Value_Event"}:
                    item[f"{metric}_ChangePct"] = post_mean / pre_mean - 1
                elif pre_mean is not None and post_mean is not None:
                    item[f"{metric}_Change"] = post_mean - pre_mean
                else:
                    item[f"{metric}_Change"] = None
            firm_changes.append(item)

        changes = pd.DataFrame(firm_changes)
        for group in ["Tanker", "Control"]:
            group_df = changes[changes["Group"].eq(group)] if not changes.empty else pd.DataFrame()
            summary_rows.append(
                {
                    "group": group,
                    "n": int(group_df["RIC"].nunique()) if not group_df.empty else 0,
                    "market_cap_change_pct": none_or_float(to_num(group_df.get("Market_Cap_Event_ChangePct", pd.Series(dtype=float))).median()),
                    "enterprise_value_change_pct": none_or_float(to_num(group_df.get("Enterprise_Value_Event_ChangePct", pd.Series(dtype=float))).median()),
                    "ev_ebitda_change": none_or_float(to_num(group_df.get("EV_EBITDA_Event_Change", pd.Series(dtype=float))).median()),
                    "pb_change": none_or_float(to_num(group_df.get("P_B_Event_Change", pd.Series(dtype=float))).median()),
                }
            )
        if len(summary_rows) == 2:
            tanker, control = summary_rows
            summary_rows.append(
                {
                    "group": "Tanker minus control",
                    "n": min(tanker["n"], control["n"]),
                    "market_cap_change_pct": none_or_float((tanker.get("market_cap_change_pct") or 0) - (control.get("market_cap_change_pct") or 0)),
                    "enterprise_value_change_pct": none_or_float((tanker.get("enterprise_value_change_pct") or 0) - (control.get("enterprise_value_change_pct") or 0)),
                    "ev_ebitda_change": none_or_float((tanker.get("ev_ebitda_change") or 0) - (control.get("ev_ebitda_change") or 0)),
                    "pb_change": none_or_float((tanker.get("pb_change") or 0) - (control.get("pb_change") or 0)),
                }
            )

    ledger_policy = {
        "current_grade": "B-reconstructed",
        "current_grade_ko": "현재 자동 생성값은 B등급 검증용입니다. 주가와 재무 스냅샷을 결합해 이벤트 전후 밸류에이션 반응을 재구성했습니다.",
        "bloomberg_grade_requirement": "Bloomberg BDH/BQL/PORT 또는 LSEG/Refinitiv에서 Date, RIC/Ticker, PX_LAST, CUR_MKT_CAP, ENTERPRISE_VALUE, EBITDA, NET_DEBT, EQY_SH_OUT, EQY_FUND_CRNCY를 같은 기준일로 추출해 업로드하면 A등급 원장으로 교체합니다.",
        "public_open_source_role": "yfinance/OpenBB/SEC/FRED는 교차검증과 누락 탐지용입니다. 유료 벤더 원장을 대체하는 확정 원장으로 표시하지 않습니다.",
        "template_file": "bloomberg_valuation_event_panel_template.csv",
    }
    return panel, summary_rows, {"missing": missing[:80], "policy": ledger_policy}


def none_or_float(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        value = float(value)
        if not math.isfinite(value):
            return None
        return value
    except Exception:
        return None


def none_or_int(value: Any) -> int | None:
    try:
        if value is None or pd.isna(value):
            return None
        return int(value)
    except Exception:
        return None


def interpretation(models: list[dict[str, Any]], car_summary: list[dict[str, Any]]) -> list[str]:
    notes = []
    base = next((m for m in models if m["name"].startswith("Base DiD")), None)
    if base:
        tp = next((c for c in base["coefficients"] if c["term"] == "Treat_Post"), None)
        if tp and tp["coef"] is not None:
            pct = tp["coef"] * 100
            sig = "통계적으로 뚜렷한" if tp.get("p_approx") is not None and tp["p_approx"] < 0.05 else "예비 표본에서는 약한"
            notes.append(
                f"Base DiD에서 Treat×Post 계수는 {pct:.2f}%p/일이며, {sig} 해운 shock(현재 사례: 홍해) 이후 탱커 주력 선사의 초과수익률 차이를 시사한다."
            )
    for row in car_summary:
        if row["difference"] is not None:
            notes.append(
                f"{row['metric']} 평균 차이는 {row['difference'] * 100:.2f}%p이다 "
                f"(탱커 n={row['n_tanker']}, 벌커/통제 n={row['n_control']})."
            )
    notes.append("공식 오류 검정상 Firm_Master Control 공식과 CAR_Calc VLOOKUP 범위 오류가 있어, 최종 논문 전 공식 수정 또는 값 고정 검증이 필요하다.")
    return notes


def write_json(payload: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def stata_safe_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        safe = clean_col(col).lower()
        safe = re.sub(r"[^a-z0-9_]+", "_", safe)
        safe = re.sub(r"_+", "_", safe).strip("_")
        if not safe:
            safe = "var"
        if safe[0].isdigit():
            safe = f"v_{safe}"
        rename[col] = safe[:32]
    return df.rename(columns=rename)


def write_stata_package(
    path: Path,
    output_dir: Path,
    valuation_panel: pd.DataFrame | None = None,
    valuation_summary: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    did = read_sheet(path, "DiD_Panel_Template")
    did["Treat_Post"] = to_num(did.get("Treat_Post", pd.Series(dtype=float)))
    if "Include" in did.columns:
        did = did[to_num(did["Include"]).eq(1)].copy()
    did["Date"] = pd.to_datetime(did["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    did_cols = [
        "Date",
        "EventDay",
        "RIC",
        "Company_Name",
        "Segment",
        "Return",
        "AR",
        "Treat",
        "Post",
        "Treat_Post",
        "Size_ln_MCap",
        "Leverage",
        "Cash_Assets",
        "Beta",
        "Market_Return",
        "Oil_Return",
        "VIX_Close",
        "BDTI_Return",
        "BCTI_Return",
        "BDI_Return",
        "Include",
    ]
    did = did[[col for col in did_cols if col in did.columns]]
    did_out = stata_safe_columns(did)
    did_csv = output_dir / "red_sea_did_panel.csv"
    did_out.to_csv(did_csv, index=False)

    car = read_sheet(path, "CAR_Calc_Template")
    if "Include" in car.columns:
        car = car[to_num(car["Include"]).eq(1)].copy()
    car_out = stata_safe_columns(car)
    car_csv = output_dir / "red_sea_car_firm_level.csv"
    car_out.to_csv(car_csv, index=False)

    event = read_sheet(path, "Event_Study_Path")
    event_out = stata_safe_columns(event)
    event_csv = output_dir / "red_sea_event_path.csv"
    event_out.to_csv(event_csv, index=False)
    valuation_csv = output_dir / "shipping_shock_valuation_event_panel.csv"
    valuation_summary_csv = output_dir / "shipping_shock_valuation_reaction_summary.csv"
    if valuation_panel is not None and not valuation_panel.empty:
        valuation_panel.to_csv(valuation_csv, index=False)
    else:
        pd.DataFrame().to_csv(valuation_csv, index=False)
    pd.DataFrame(valuation_summary or []).to_csv(valuation_summary_csv, index=False)

    do_text = '''/**********************************************************************
 Red Sea shipping shock regression package
 Generated from: attached workbook export
 Purpose: university thesis-ready DiD/event-study workflow

 Core interpretation:
   treat_post coefficient = incremental post-shock reaction of tanker-focused
   listed shipping firms relative to dry-bulk/control firms.

 Before thesis submission:
   1. Confirm source labels: Bloomberg vs Refinitiv/LSEG vs Clarksons.
   2. Fix workbook formula issues flagged in the app/Data_Audit.
   3. Re-run this do-file after sample is frozen.
**********************************************************************/

clear all
set more off
version 17

capture confirm file "red_sea_did_panel.csv"
if _rc {
    di as error "Run this do-file from inside the red_sea_stata_package folder."
    exit 601
}

capture log close
log using "red_sea_stata_run.log", replace text

* Optional open-source/user-written packages.
* reghdfe: high-dimensional fixed effects; estout/esttab: publication tables.
capture which reghdfe
if _rc ssc install reghdfe, replace
capture which esttab
if _rc ssc install estout, replace

import delimited "red_sea_did_panel.csv", clear varnames(1)

gen date_stata = date(date, "YMD")
format date_stata %td
encode ric, gen(firm_id)
xtset firm_id date_stata

label var ar "Daily abnormal return"
label var return "Daily stock return"
label var treat "Tanker-focused firm"
label var post "Post shipping shock (current case: Red Sea)"
label var treat_post "Treat x Post"
label var market_return "Market return"
label var oil_return "Brent/oil return"
label var vix_close "VIX close"
label var bdti_return "Baltic Dirty Tanker Index return"
label var bcti_return "Baltic Clean Tanker Index return"
label var bdi_return "Baltic Dry Index return"

eststo clear

* 1. Baseline DiD.
reg ar treat post treat_post, vce(cluster firm_id)
eststo m1

* 2. Controls: firm fundamentals and macro/market context.
reg ar treat post treat_post size_ln_mcap leverage market_return oil_return vix_close, vce(cluster firm_id)
eststo m2

* 3. Freight-index sensitivity.
reg ar treat post treat_post bdti_return bcti_return bdi_return, vce(cluster firm_id)
eststo m3

* 4. Two-way fixed effects, preferred if reghdfe works.
capture noisily reghdfe ar treat_post market_return oil_return vix_close bdti_return bcti_return bdi_return, absorb(firm_id eventday) vce(cluster firm_id)
if !_rc eststo m4

* Built-in fallback FE model.
reg ar treat_post market_return oil_return vix_close bdti_return bcti_return bdi_return i.firm_id i.eventday, vce(cluster firm_id)
eststo m5

* Raw return robustness.
reg return treat post treat_post market_return oil_return vix_close, vce(cluster firm_id)
eststo m6

capture noisily esttab m1 m2 m3 m4 m5 m6 using "red_sea_regression_table.rtf", replace ///
    b(4) se(4) r2 ar2 label star(* 0.10 ** 0.05 *** 0.01) ///
    title("Red Sea Shock: Tanker vs Dry Bulk Listed Shipping Firms") ///
    keep(treat post treat_post market_return oil_return vix_close bdti_return bcti_return bdi_return size_ln_mcap leverage)

capture noisily esttab m1 m2 m3 m4 m5 m6 using "red_sea_regression_table.csv", replace ///
    b(6) se(6) r2 label star(* 0.10 ** 0.05 *** 0.01)

* Event-study path chart data is separate; import for quick graph.
preserve
import delimited "red_sea_event_path.csv", clear varnames(1)
twoway line car_diff_t_c eventday, sort yline(0) xline(0) ///
    title("CAR Difference: Tanker minus Control") ///
    xtitle("Event day") ytitle("CAR difference")
graph export "red_sea_event_path.png", replace width(1800)
restore

* Valuation reaction bridge.
* This is a reconstructed validation panel. For thesis-final valuation
* reactions, replace the CSV with Bloomberg/LSEG/Refinitiv date-stamped export.
preserve
import delimited "shipping_shock_valuation_event_panel.csv", clear varnames(1)
capture confirm variable ev_ebitda_event
if !_rc {
    gen date_stata = date(date, "YMD")
    format date_stata %td
    encode ric, gen(val_firm_id)
    reg ev_ebitda_event treat post treat_post, vce(cluster val_firm_id)
    estimates store v1
    reg p_b_event treat post treat_post, vce(cluster val_firm_id)
    estimates store v2
    capture noisily esttab v1 v2 using "shipping_shock_valuation_table.csv", replace ///
        b(6) se(6) r2 label star(* 0.10 ** 0.05 *** 0.01)
}
restore

log close
'''
    do_file = output_dir / "red_sea_regression.do"
    do_file.write_text(do_text, encoding="utf-8")
    run_script = output_dir / "run_red_sea_stata.sh"
    if run_script.exists():
        run_script.unlink()
    bat_file = output_dir / "run_red_sea_stata_windows.bat"
    bat_file.write_text(
        r'''@echo off
setlocal
cd /d "%~dp0"

set "DO_FILE=%CD%\red_sea_regression.do"
set "STATA_BIN="

for %%P in (
  "C:\Program Files\Stata18\StataMP-64.exe"
  "C:\Program Files\Stata18\StataSE-64.exe"
  "C:\Program Files\Stata18\StataBE-64.exe"
  "C:\Program Files\Stata17\StataMP-64.exe"
  "C:\Program Files\Stata17\StataSE-64.exe"
  "C:\Program Files\Stata17\StataBE-64.exe"
  "C:\Program Files\Stata16\StataMP-64.exe"
  "C:\Program Files\Stata16\StataSE-64.exe"
  "C:\Program Files\Stata16\StataIC-64.exe"
) do (
  if exist %%~P (
    set "STATA_BIN=%%~P"
    goto :run
  )
)

where stata-mp >nul 2>nul
if %errorlevel%==0 (
  set "STATA_BIN=stata-mp"
  goto :run
)
where stata-se >nul 2>nul
if %errorlevel%==0 (
  set "STATA_BIN=stata-se"
  goto :run
)
where stata >nul 2>nul
if %errorlevel%==0 (
  set "STATA_BIN=stata"
  goto :run
)

echo Stata executable was not found.
echo.
echo Option 1: Open Stata manually, then open red_sea_regression.do and run all.
echo Option 2: Edit this .bat file and set STATA_BIN to your Stata exe path.
echo.
pause
exit /b 127

:run
echo Using Stata: %STATA_BIN%
"%STATA_BIN%" /e do "%DO_FILE%"
echo Done. Check red_sea_stata_run.log and red_sea_regression_table files.
pause
''',
        encoding="utf-8",
    )
    how_to_file = output_dir / "README_KR_HOW_TO_USE.md"
    how_to_file.write_text(
        f"""# 해운 Shock Stata 패키지 사용방법 (현재 사례: 홍해)

## 1. 동생에게 보내야 하는 것

### 보기만 할 때
- GitHub Pages 앱 링크
- 앱에서는 데이터 감사 결과, 예비 회귀 계수, CAR 차이, 밸류에이션 반응, 논문 초안, Stata 실행 안내를 볼 수 있습니다.

### Stata로 직접 회귀를 돌릴 때
아래 폴더 전체를 압축해서 보내세요.

`red_sea_stata_package`

폴더 안 파일:
- `red_sea_did_panel.csv`: Stata 회귀용 패널 데이터
- `red_sea_car_firm_level.csv`: 회사별 CAR 요약 데이터
- `red_sea_event_path.csv`: 이벤트 스터디 경로 데이터
- `shipping_shock_valuation_event_panel.csv`: 이벤트 전후 Market Cap/EV/EV/EBITDA 재구성 패널
- `shipping_shock_valuation_reaction_summary.csv`: 탱커/통제군 밸류에이션 반응 요약
- `red_sea_regression.do`: Stata 회귀 명령 파일
- `run_red_sea_stata_windows.bat`: Windows/LG 노트북 실행 파일
- `red_sea_thesis_draft.md`: 논문 초안 파일
- `README_KR_HOW_TO_USE.md`: 이 사용방법 파일

주의: Bloomberg, Refinitiv/LSEG, Clarksons 같은 라이선스 원자료가 섞일 수 있으므로 이 폴더는 공개 GitHub에 올리지 말고 동생에게만 따로 보내세요.

## 2. Windows / LG 노트북에서 실행하는 방법

1. 받은 압축파일을 풉니다.
2. 폴더 안의 `run_red_sea_stata_windows.bat`를 더블클릭합니다.
3. Stata가 자동으로 열리면 기다립니다.
4. 실행이 끝나면 같은 폴더에 아래 결과물이 생깁니다.
   - `red_sea_stata_run.log`
   - `red_sea_regression_table.rtf`
   - `red_sea_regression_table.csv`
   - `red_sea_event_path.png`

자동 실행이 안 되면:
1. Stata를 직접 엽니다.
2. `red_sea_regression.do` 파일을 엽니다.
3. 전체 실행합니다.

## 3. 결과를 어떻게 읽는가

가장 중요한 값은 `Treat x Post` 또는 `treat_post`입니다.

- 양수: 해운 shock 이후 탱커 주력 선사가 벌커 통제군보다 더 강하게 상승했다는 방향
- 음수: 해운 shock 이후 탱커 주력 선사가 벌커 통제군보다 상대적으로 약했다는 방향
- p-value < 0.05: 통계적으로 비교적 강한 결과
- p-value < 0.10: 논문에서 약한 유의성으로 조심스럽게 언급 가능
- p-value가 크면: 방향성은 참고하되 결론을 강하게 쓰면 안 됩니다.

## 4. 논문에 쓰기 전 필수 확인

1. 원자료 출처를 확정하세요.
   - 현재 엑셀에는 Refinitiv/LSEG로 표기된 부분이 있습니다.
   - Bloomberg에서 뽑은 자료라면 Bloomberg 추출 로그 또는 스크린샷을 따로 보관하세요.
   - Market Cap/EV/EBITDA 반응은 Bloomberg/LSEG/Refinitiv 날짜별 원장으로 교체하면 A등급 확정값이 됩니다.
2. 앱의 데이터 감사에서 표시된 공식 오류를 먼저 수정하거나, Stata 패키지 생성값이 수정된 기준인지 확인하세요.
3. 최종 표는 Stata 결과표를 기준으로 쓰세요.
4. 다른 논문 문장을 복사하지 말고, 이 데이터의 표본·이벤트 날짜·회귀 결과·한계를 중심으로 직접 작성하세요.
""",
        encoding="utf-8",
    )
    return {
        "directory": "red_sea_stata_package",
        "files": [
            did_csv.name,
            car_csv.name,
            event_csv.name,
            valuation_csv.name,
            valuation_summary_csv.name,
            do_file.name,
            bat_file.name,
            how_to_file.name,
        ],
        "windows_run_command": "run_red_sea_stata_windows.bat",
        "note": "Windows/LG 노트북 기준: red_sea_stata_package 폴더를 받은 뒤 run_red_sea_stata_windows.bat를 더블클릭하세요. 자동 실행이 안 되면 Stata에서 red_sea_regression.do를 직접 열어 실행합니다.",
    }


def append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=min(ws.max_row, 200)):
            for item in cell:
                if item.value is not None:
                    max_len = max(max_len, min(60, len(str(item.value))))
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(max_len + 2, 55))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_enhanced_workbook(payload: dict[str, Any], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    append_rows(
        ws,
        [
            ["Section", "Value", "Note"],
            ["Purpose", "Red Sea shock DiD audit and regression summary", "Generated from attached workbook; raw vendor extract is not overwritten."],
            ["Input workbook", payload["summary"]["input_workbook"], ""],
            ["Generated at", payload["summary"]["generated_at"], ""],
            ["Main caution", "Formula defects found", "Review Data_Audit before using regression output in thesis."],
        ],
    )

    ws = wb.create_sheet("Data_Audit")
    append_rows(ws, [["Severity", "Check", "Status", "Affected", "Detail", "Fix"]])
    for item in payload["checks"]:
        append_rows(ws, [[item["severity"], item["check"], item["status"], item["affected"], item["detail"], item["fix"]]])

    ws = wb.create_sheet("Regression_Results")
    append_rows(ws, [["Model", "Outcome", "N", "R2", "Term", "Coef", "SE_HC1", "t", "p_approx", "Interpretation"]])
    for model in payload["regressions"]:
        for coef in model["coefficients"]:
            interp = ""
            if coef["term"] == "Treat_Post" and coef["coef"] is not None:
                interp = f"Incremental post-shock effect: {coef['coef'] * 100:.2f}%p/day"
            append_rows(
                ws,
                [
                    [
                        model["name"],
                        model["outcome"],
                        model["n"],
                        model["r2"],
                        coef["term"],
                        coef["coef"],
                        coef["se_hc1"],
                        coef["t"],
                        coef["p_approx"],
                        interp,
                    ]
                ],
            )

    ws = wb.create_sheet("CAR_Summary")
    append_rows(ws, [["Metric", "Tanker Avg", "Control Avg", "Difference", "N Tanker", "N Control"]])
    for row in payload["car_summary"]:
        append_rows(ws, [[row["metric"], row["tanker_avg"], row["control_avg"], row["difference"], row["n_tanker"], row["n_control"]]])

    ws = wb.create_sheet("Event_Path")
    append_rows(ws, [["EventDay", "AR_Treat", "AR_Control", "AR_Diff", "CAR_Treat", "CAR_Control", "CAR_Diff", "N_Treat", "N_Control"]])
    for row in payload["event_path"]:
        append_rows(
            ws,
            [[row["event_day"], row["ar_treat"], row["ar_control"], row["ar_diff"], row["car_treat"], row["car_control"], row["car_diff"], row["n_treat"], row["n_control"]]],
        )

    ws = wb.create_sheet("Source_Map")
    append_rows(ws, [["Item", "Workbook Sheet", "Workbook Source", "Validation Source", "Status", "Note"]])
    for row in payload["source_map"]:
        append_rows(ws, [[row["item"], row["workbook_sheet"], row["workbook_source"], row["validation_source"], row["status"], row["note"]]])

    for ws in wb.worksheets:
        style_sheet(ws)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def reset_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_completed_source_workbook(
    input_path: Path,
    payload: dict[str, Any],
    output: Path,
    valuation_panel_df: pd.DataFrame | None = None,
) -> dict[str, Any]:
    """Create a completed copy of the user's original modeling workbook."""
    wb = load_workbook(input_path)

    # Mark workbook for recalculation when opened in Excel/Stata user's machine.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    if "Firm_Master" in wb.sheetnames:
        ws = wb["Firm_Master"]
        for row in range(4, ws.max_row + 1):
            if ws.cell(row, 3).value in (None, ""):
                continue
            ws.cell(row, 8).value = f'=IF(F{row}>=60,1,0)'
            ws.cell(row, 9).value = f'=IF(AND(G{row}>=70,F{row}<=20),1,0)'
            ws.cell(row, 10).value = f'=IF(OR(H{row}=1,I{row}=1),1,0)'
            ws.cell(row, 11).value = f'=IFERROR(INDEX(Financials_Raw!$L:$L,MATCH($C{row},Financials_Raw!$A:$A,0)),"")'
            ws.cell(row, 12).value = f'=IFERROR(INDEX(Financials_Raw!$M:$M,MATCH($C{row},Financials_Raw!$A:$A,0)),"")'
            ws.cell(row, 13).value = f'=IFERROR(INDEX(Financials_Raw!$N:$N,MATCH($C{row},Financials_Raw!$A:$A,0)),"")'

    if "CAR_Calc_Template" in wb.sheetnames:
        ws = wb["CAR_Calc_Template"]
        for row in range(4, ws.max_row + 1):
            if ws.cell(row, 1).value in (None, ""):
                continue
            ws.cell(row, 9).value = f'=IFERROR(INDEX(Firm_Master!$N:$N,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 10).value = f'=IFERROR(INDEX(Firm_Master!$O:$O,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 11).value = f'=IFERROR(INDEX(Firm_Master!$P:$P,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 12).value = f'=IFERROR(INDEX(Firm_Master!$Q:$Q,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 13).value = f'=IFERROR(INDEX(Firm_Master!$K:$K,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 14).value = f'=IFERROR(INDEX(Firm_Master!$L:$L,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 15).value = f'=IFERROR(INDEX(Firm_Master!$M:$M,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 17).value = f'=IFERROR(INDEX(Firm_Master!$J:$J,MATCH($A{row},Firm_Master!$C:$C,0)),"")'
            ws.cell(row, 18).value = f'=IFERROR(INDEX(Firm_Master!$R:$R,MATCH($A{row},Firm_Master!$C:$C,0)),"")'

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        start = ws.max_row + 2
        rows = [
            ["Completed workbook generated", payload["summary"]["generated_at"], "Formula fixes and regression/audit sheets added", "Use this file for thesis modeling", "Completed"],
            ["Firm_Master Control formulas", "Fixed", "Control now references same row G/F values", "Recalculate in Excel", "Completed"],
            ["CAR metadata formulas", "Fixed", "Alpha/Beta/R2/N/Size/Leverage/Cash now use INDEX/MATCH by RIC", "Review after Excel recalculation", "Completed"],
            ["Stata package", payload["stata_package"]["directory"], "Windows/LG launcher generated", "Send folder privately", "Completed"],
        ]
        for i, values in enumerate(rows, start=start):
            for col, value in enumerate(values, start=1):
                ws.cell(i, col).value = value

    if valuation_panel_df is None:
        valuation_panel_df = pd.DataFrame()

    ws = reset_sheet(wb, "Completion_README")
    append_rows(
        ws,
        [
            ["항목", "내용", "사용방법"],
            ["이 파일", "사용자가 준 원본 모델링 엑셀을 복사해 완성한 버전", "원본은 보존하고 이 completed 파일로 분석하세요."],
            ["수정 1", "Firm_Master Control 공식 행 참조 오류 수정", "Control은 DryBulk_%>=70 및 Tanker_%<=20 기준입니다."],
            ["수정 2", "CAR_Calc_Template의 Alpha/Beta/R2/N/Size/Leverage/Cash 누락 연결", "RIC 기준 INDEX/MATCH로 Firm_Master 값을 가져오게 했습니다."],
            ["추가 1", "Data_Audit_Fixes", "논문 전 반드시 확인할 데이터 오류와 수정 지시입니다."],
            ["추가 2", "Regression_Results", "Python 예비 DiD 결과입니다. 최종 논문 표는 Stata 재실행 결과로 확정하세요."],
            ["추가 3", "Valuation_Reaction", "Market Cap/EV/EV-EBITDA/P-B 이벤트 전후 반응을 검증용으로 계산했습니다."],
            ["추가 4", "Bloomberg_Grade_Ledger", "Bloomberg/LSEG/Refinitiv 최종 원장으로 교체할 때 필요한 필드와 정확도 등급입니다."],
            ["추가 5", "Stata_Instructions", "Windows/LG 노트북에서 실행하는 방법입니다."],
            ["주의", "라이선스 데이터 포함 가능", "Bloomberg/Refinitiv/LSEG/Clarksons 원자료는 공개 GitHub에 올리지 마세요."],
        ],
    )

    ws = reset_sheet(wb, "Data_Audit_Fixes")
    append_rows(ws, [["Severity", "한글 점검", "한글 해석", "한글 조치", "English Check", "Status", "Affected", "Detail", "Fix", "Completed_File_Action"]])
    for item in payload["checks"]:
        action = ""
        if item["check"] == "Control group formula row reference":
            action = "Firm_Master!I rows now reference same-row F/G cells."
        elif item["check"] == "CAR VLOOKUP index range":
            action = "CAR_Calc_Template lookup formulas replaced by INDEX/MATCH."
        else:
            action = "Review required before thesis final submission."
        append_rows(
            ws,
            [[
                item["severity"],
                item.get("check_ko", item["check"]),
                item.get("detail_ko", item["detail"]),
                item.get("fix_ko", item["fix"]),
                item["check"],
                item["status"],
                item["affected"],
                item["detail"],
                item["fix"],
                action,
            ]],
        )

    ws = reset_sheet(wb, "Regression_Results")
    append_rows(ws, [["Model", "Outcome", "N", "R2", "Term", "Coef", "SE_HC1", "t", "p_approx", "Use_in_thesis"]])
    for model in payload["regressions"]:
        for coef in model["coefficients"]:
            use = "Core DiD coefficient" if coef["term"] == "Treat_Post" else "Control / diagnostic"
            append_rows(ws, [[model["name"], model["outcome"], model["n"], model["r2"], coef["term"], coef["coef"], coef["se_hc1"], coef["t"], coef["p_approx"], use]])

    ws = reset_sheet(wb, "CAR_Summary_Completed")
    append_rows(ws, [["Metric", "Tanker Avg", "Control Avg", "Difference", "N Tanker", "N Control", "Interpretation"]])
    for row in payload["car_summary"]:
        append_rows(
            ws,
            [[row["metric"], row["tanker_avg"], row["control_avg"], row["difference"], row["n_tanker"], row["n_control"], "Tanker minus control CAR difference"]],
        )

    ws = reset_sheet(wb, "Valuation_Reaction")
    append_rows(
        ws,
        [["Group", "N", "Market Cap Change %", "EV Change %", "EV/EBITDA Change", "P/B Change", "한글 해석"]],
    )
    for row in payload.get("valuation_reaction", {}).get("summary", []):
        label = row.get("group", "")
        explanation = {
            "Tanker": "탱커 주력 선사의 이벤트 전후 평균 변화입니다.",
            "Control": "벌커/통제군 선사의 이벤트 전후 평균 변화입니다.",
            "Tanker minus control": "탱커 반응에서 통제군 반응을 뺀 차이입니다.",
        }.get(label, "")
        append_rows(
            ws,
            [[
                label,
                row.get("n"),
                row.get("market_cap_change_pct"),
                row.get("enterprise_value_change_pct"),
                row.get("ev_ebitda_change"),
                row.get("pb_change"),
                explanation,
            ]],
        )

    ws = reset_sheet(wb, "Valuation_Event_Panel")
    if not valuation_panel_df.empty:
        append_rows(ws, [list(valuation_panel_df.columns)])
        for _, row in valuation_panel_df.iterrows():
            append_rows(ws, [[none_or_float(v) if isinstance(v, (int, float, np.floating)) else v for v in row.tolist()]])
    else:
        append_rows(ws, [["Message"], ["Valuation event panel could not be generated from the attached workbook."]])

    ws = reset_sheet(wb, "Bloomberg_Grade_Ledger")
    policy = payload.get("valuation_reaction", {}).get("meta", {}).get("policy", {})
    append_rows(
        ws,
        [
            ["항목", "내용"],
            ["현재 정확도 등급", policy.get("current_grade_ko", "")],
            ["A등급 확정 원장 조건", policy.get("bloomberg_grade_requirement", "")],
            ["공개 오픈소스 역할", policy.get("public_open_source_role", "")],
            ["업로드 템플릿", policy.get("template_file", "")],
            ["필수 필드", "Date, RIC, Company_Name, PX_LAST, CUR_MKT_CAP, ENTERPRISE_VALUE, EBITDA, NET_DEBT, EQY_SH_OUT, EQY_FUND_CRNCY, Source, Source_Timestamp"],
            ["원칙", "Bloomberg/Clarksons/LSEG 원장은 공개 GitHub에 올리지 말고, 앱에는 익명화 또는 요약값만 올립니다."],
        ],
    )

    ws = reset_sheet(wb, "Source_Map")
    append_rows(ws, [["Item", "Workbook Sheet", "Workbook Source", "Validation Source", "Status", "Note"]])
    for row in payload["source_map"]:
        append_rows(ws, [[row["item"], row["workbook_sheet"], row["workbook_source"], row["validation_source"], row["status"], row["note"]]])

    ws = reset_sheet(wb, "Stata_Instructions")
    append_rows(
        ws,
        [
            ["항목", "내용"],
            ["Stata 패키지 폴더", "red_sea_stata_package"],
            ["Windows/LG 실행", payload["stata_package"]["windows_run_command"]],
            ["공통 do-file", "red_sea_regression.do"],
            ["결과 파일", "red_sea_regression_table.rtf / red_sea_regression_table.csv / red_sea_event_path.png / red_sea_stata_run.log"],
            ["중요 계수", "Treat_Post: 해운 shock 이후 탱커 주력 선사의 벌커 통제군 대비 추가 반응"],
            ["밸류에이션 패널", "shipping_shock_valuation_event_panel.csv를 Bloomberg/LSEG 원장으로 교체하면 EV/EBITDA 회귀도 같은 do-file에서 실행됩니다."],
            ["주의", "Stata가 설치된 Windows/LG 노트북에서 bat 파일 또는 do-file을 실행하세요."],
            ["동생에게 전달", "이 completed 엑셀 + red_sea_stata_package 폴더 전체"],
        ],
    )

    ws = reset_sheet(wb, "Thesis_Draft_Notes")
    append_rows(ws, [["Section", "Draft Note"]])
    append_rows(ws, [["Research question", payload["originality"]["thesis_angle"]]])
    append_rows(ws, [["Research gap", payload["originality"]["research_gap"]]])
    append_rows(ws, [["Non-overlap rule", payload["originality"]["non_overlap_rule"]]])
    for note in payload["interpretation"]:
        append_rows(ws, [["Preliminary interpretation", note]])
    for guard in payload["originality"]["plagiarism_guard"]:
        append_rows(ws, [["Plagiarism guard", guard]])

    for sheet_name in [
        "Completion_README",
        "Data_Audit_Fixes",
        "Regression_Results",
        "CAR_Summary_Completed",
        "Valuation_Reaction",
        "Valuation_Event_Panel",
        "Bloomberg_Grade_Ledger",
        "Source_Map",
        "Stata_Instructions",
        "Thesis_Draft_Notes",
    ]:
        style_sheet(wb[sheet_name])

    # Make the most important sheets easy to open first.
    wb.active = wb.sheetnames.index("Completion_README")
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "path": output.name,
        "actions": [
            "Fixed Firm_Master control formulas",
            "Filled CAR_Calc_Template metadata formulas",
            "Added audit, regression, valuation, source, Stata, and thesis-note sheets",
        ],
    }


def coef_from_payload(payload: dict[str, Any], model_index: int, term: str) -> dict[str, Any] | None:
    try:
        return next(
            item
            for item in payload["regressions"][model_index]["coefficients"]
            if item["term"] == term
        )
    except Exception:
        return None


def fmt_pct_point(value: Any) -> str:
    if value is None:
        return "-"
    try:
        return f"{float(value) * 100:+.2f}%p"
    except Exception:
        return "-"


def write_thesis_draft_file(payload: dict[str, Any], output_dir: Path) -> dict[str, Any]:
    base = coef_from_payload(payload, 0, "Treat_Post")
    controlled = coef_from_payload(payload, 1, "Treat_Post")
    freight = coef_from_payload(payload, 2, "BDI_Return")
    checks = payload.get("checks", [])
    high_checks = [item for item in checks if item.get("severity") == "High"]
    car_lines = []
    for row in payload.get("car_summary", []):
        car_lines.append(
            f"| {row['metric']} | {fmt_pct_point(row['tanker_avg'])} | {fmt_pct_point(row['control_avg'])} | {fmt_pct_point(row['difference'])} |"
        )
    regression_lines = []
    for model in payload.get("regressions", []):
        item = next((coef for coef in model.get("coefficients", []) if coef.get("term") == "Treat_Post"), None)
        regression_lines.append(
            f"| {model['name']} | {fmt_pct_point(item.get('coef') if item else None)} | {item.get('t') if item else '-'} | {item.get('p_approx') if item else '-'} |"
        )
    valuation_rows = payload.get("valuation_reaction", {}).get("summary", [])
    valuation_lines = []
    for row in valuation_rows:
        valuation_lines.append(
            f"| {row.get('group')} | {row.get('n')} | {fmt_pct_point(row.get('market_cap_change_pct'))} | {fmt_pct_point(row.get('enterprise_value_change_pct'))} | {row.get('ev_ebitda_change')} | {row.get('pb_change')} |"
        )
    valuation_policy = payload.get("valuation_reaction", {}).get("meta", {}).get("policy", {})
    text = f"""# 해운 Shock과 상장 탱커·벌커 선사의 시장·밸류에이션 반응 초안

## 초록

본 연구의 목적은 해운 shock이 상장 탱커 주력 선사와 벌커 주력 선사의 주가, 초과수익률, 기업가치평가에 미치는 차별적 반응을 검정하는 것이다. 현재 실증 사례는 2023년 11월 19일 Galaxy Leader 나포와 2023년 11월 20일 첫 거래일 반응으로 정의한 홍해 공급망 shock이다. 분석 표본은 {payload['summary']['included_firms']}개 포함 회사와 {payload['summary']['did_rows']:,}개의 회귀 패널 관측치로 구성된다. 핵심 DiD 계수인 Treat×Post는 Base 모형에서 {fmt_pct_point(base.get('coef') if base else None)}, 통제변수 포함 모형에서 {fmt_pct_point(controlled.get('coef') if controlled else None)}로 추정되었다. 다만 원본 데이터 감사에서 높은 등급 오류 유형 {len(high_checks)}건이 확인되어, 최종 결론은 completed 엑셀과 Stata 재실행 결과로 확정한다.

## 1. 연구 질문

{payload['originality']['thesis_angle']}

## 2. 연구의 차별성

{payload['originality']['research_gap']}

본 논문은 기존 문헌 문장을 복제하지 않고, 직접 구성한 상장사 표본, 선대 노출도, 해운 shock 이벤트 정의, 데이터 감사 로그, Stata 회귀 결과를 중심으로 독립적으로 서술한다.

## 3. 데이터와 출처

- 원본 모델링 파일: Red_Sea_DiD_Model_drafting.xlsx
- 완성본 모델링 파일: Red_Sea_DiD_Model_completed.xlsx
- Stata 실행 폴더: red_sea_stata_package
- 가격 데이터 기간: {payload['summary']['price_date_min']} ~ {payload['summary']['price_date_max']}
- 주가 관측치: {payload['summary']['price_observations']:,}
- DiD 패널 관측치: {payload['summary']['did_rows']:,}
- 패널 회사 수: {payload['summary']['did_firms']}

현재 엑셀에는 Refinitiv/LSEG Workspace 표기가 존재한다. Bloomberg에서 추출한 값이라면 Bloomberg 추출 로그 또는 스크린샷을 별도 보관하고, 최종 논문에서는 하나의 출처 체계로 통일한다.

## 4. 방법론

기본 모형은 다음과 같다.

`AR_it = alpha + beta1 Treat_i + beta2 Post_t + beta3 Treat_i x Post_t + Controls_it + error_it`

여기서 `Treat_i`는 탱커 주력 선사, `Post_t`는 해운 shock 이후 기간, `Treat_i x Post_t`는 shock 이후 탱커 주력 선사의 벌커 통제군 대비 추가 반응을 의미한다.

## 5. 예비 회귀 결과

| 모형 | Treat×Post 계수 | t | p-value |
| --- | --- | --- | --- |
{chr(10).join(regression_lines)}

운임 민감도 모형의 BDI_Return 계수는 {freight.get('coef') if freight else '-'}이며, 이는 벌크 운임 환경을 별도 통제해야 함을 시사한다.

## 6. 기업가치평가 반응

현재 자동 계산값은 {valuation_policy.get('current_grade_ko', 'B등급 검증용 재구성 패널')}으로 분류한다. 논문 최종본에서는 {valuation_policy.get('bloomberg_grade_requirement', 'Bloomberg/LSEG/Refinitiv 날짜별 원장')}을 사용해 Market Cap, EV, EBITDA, P/B를 같은 기준일로 맞춘다.

| 그룹 | N | Market Cap 변화율 | EV 변화율 | EV/EBITDA 변화 | P/B 변화 |
| --- | --- | --- | --- | --- | --- |
{chr(10).join(valuation_lines)}

## 7. CAR 결과

| 창 | 탱커 평균 | 통제 평균 | 차이 |
| --- | --- | --- | --- |
{chr(10).join(car_lines)}

## 8. 데이터 감사 결과

{chr(10).join(f"- [{item['severity']}] {item.get('check_ko', item['check'])}: {item.get('detail_ko', item['detail'])}" for item in checks)}

## 9. 최종 제출 전 할 일

1. completed 엑셀을 열어 공식이 재계산되는지 확인한다.
2. Windows/LG 노트북에서 `run_red_sea_stata_windows.bat`를 실행한다.
3. Stata 결과표의 Treat×Post 계수, 표준오차, p-value를 최종 논문 표에 사용한다.
4. Market Cap/EV/EBITDA는 Bloomberg/LSEG/Refinitiv 날짜별 원장 또는 공시 기반 값으로 교체한다.
5. 데이터 출처를 Bloomberg 또는 Refinitiv/LSEG 중 하나로 명확히 통일한다.
6. 다른 논문 문장을 복사하지 않고, 본 표본과 본 회귀 결과 중심으로 서술한다.
"""
    output = output_dir / "red_sea_thesis_draft.md"
    output.write_text(text, encoding="utf-8")
    return {"path": output.name, "note": "논문 초안은 대시보드에서 바로 볼 수 있고, Stata 패키지에는 백업 파일로도 생성됩니다."}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--completed-workbook", type=Path, default=DEFAULT_COMPLETED_WORKBOOK)
    parser.add_argument("--stata-dir", type=Path, default=DEFAULT_STATA_DIR)
    args = parser.parse_args()

    summary, checks = workbook_audit(args.input)
    regressions, event_path, car_summary = run_models(args.input)
    valuation_panel, valuation_summary, valuation_meta = build_valuation_event_panel(args.input)
    stata_package = write_stata_package(args.input, args.stata_dir, valuation_panel, valuation_summary)
    payload = {
        "summary": summary,
        "checks": checks,
        "regressions": regressions,
        "event_path": event_path,
        "car_summary": car_summary,
        "source_map": SOURCE_MAP,
        "stata_package": stata_package,
        "valuation_reaction": {
            "summary": valuation_summary,
            "meta": valuation_meta,
            "panel_rows": int(len(valuation_panel)),
            "panel_firms": int(valuation_panel["RIC"].nunique()) if not valuation_panel.empty else 0,
            "event_window": "[-20,+20] valuation reaction summary; full CSV covers regression event-window dates available in the workbook.",
        },
        "originality": {
            "research_gap": "기존 해운 event study는 운임지수 또는 해운주 전반의 반응을 보는 경우가 많다. 본 연구는 특정 해운 shock을 사례로 삼되, 상장 탱커 주력 선사와 벌커 주력 선사를 분리하고, 선대 노출도·해운지수·경제지수·기업가치평가/초과수익률 반응을 연결한다.",
            "non_overlap_rule": "문헌의 문장을 복제하지 않고, 본 데이터셋의 표본 정의·검정 로그·Stata 회귀 결과·선대 노출도 기준을 중심으로 독자적 서술을 작성한다.",
            "thesis_angle": "해운 shock은 상장 탱커 주력 선사와 벌커 주력 선사의 주가·초과수익률·기업가치평가 민감도에 서로 다른 영향을 주는가? 현재 실증 사례는 홍해 공급망 shock으로 설정한다.",
            "plagiarism_guard": [
                "문헌은 주장 근거와 방법론 위치를 인용하는 용도로만 사용한다.",
                "초록·서론·결론은 앱의 실제 표본 수, 이벤트 날짜, 회귀 계수, 데이터 한계를 반영해 매번 새로 생성한다.",
                "출처 없는 수치와 타 논문 문장 구조를 그대로 쓰지 않는다.",
            ],
        },
        "interpretation": interpretation(regressions, car_summary),
        "methodology": {
            "event": "Red Sea Crisis / Galaxy Leader seizure. Trigger date 2023-11-19; first trading reaction date 2023-11-20.",
            "design": "Difference-in-differences panel around event window [-60,+60] trading days.",
            "treatment": "Tanker-focused listed shipping firms. Control group is dry-bulk-focused listed shipping firms.",
            "outcomes": ["Daily abnormal return from market model", "Daily raw return", "CAR windows [-1,+5], [-1,+10], [-1,+20]", "Market Cap/EV/EV-EBITDA/P-B event-window reaction when vendor valuation panel is available"],
            "core_coefficient": "Treat×Post measures incremental post-shock reaction of tanker firms relative to dry-bulk/control firms.",
            "caution": "This is a vendor-extract audit and preliminary regression run. Thesis-final inference should use corrected formulas, reconciled source labels, Bloomberg/LSEG/Refinitiv date-stamped valuation ledger for Market Cap/EV/EBITDA, and Stata/Python/R rerun with clustered standard errors.",
        },
    }
    completed = write_completed_source_workbook(args.input, payload, args.completed_workbook, valuation_panel)
    payload["completed_workbook"] = completed
    thesis_draft = write_thesis_draft_file(payload, args.stata_dir)
    payload["thesis_draft"] = thesis_draft
    if thesis_draft["path"] not in payload["stata_package"]["files"]:
        payload["stata_package"]["files"].append(thesis_draft["path"])
    write_json(payload, args.json)
    write_enhanced_workbook(payload, args.workbook)
    print(f"Wrote {args.json}")
    print(f"Wrote {args.workbook}")
    print(f"Wrote {args.completed_workbook}")


if __name__ == "__main__":
    main()
